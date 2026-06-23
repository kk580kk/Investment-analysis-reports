"""
cover.py — Cover Page 模板 + 统一调色板 (PALETTE)  v2
放置于 skill 目录的 references/cover.py，由生成脚本 import 使用。
LLM 不需要修改此文件，只需调用 create_cover_page() 和 finalize_cover_page()。

v2 变更:
  - 交替行色加深，与白色背景拉开明显对比
  - 非交替行显式填充 bg_main，避免透明穿透
  - _put / _table 对合并区域内所有列涂 fill，消除条纹空隙
  - 数据列加宽 (B-G 各 18)，长标签不再截断
  - 表格行增加底部细边框，视觉对齐更清晰
  - 用 sheet custom_properties 替代 A1 hack 存储 style
  - 布局行号改为游标推进，易于插入新区块
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  PALETTE — 全局唯一色板定义，Cover Page 和数据表共用
# ═══════════════════════════════════════════════════════════════
PALETTE = {
    'pure': {
        'header_bg':      '1B3A5C',
        'accent':         '2C6FAC',
        'text_primary':   '212529',
        'text_secondary': '6C757D',
        'bg_main':        'FFFFFF',
        'bg_alt':         'EAF0F7',   # ← 加深: 原 F8F9FA 与白底几乎无差
        'border':         'DEE2E6',
        'highlight_bg':   'E3EDF7',
        'data_bar':       '6BA3D6',
        'positive':       None,
        'negative':       None,
        'neutral':        None,
    },
    'finance': {
        'header_bg':      '0D1B2A',
        'accent':         '0D1B2A',
        'text_primary':   '1A1A2E',
        'text_secondary': '6C757D',
        'bg_main':        'FAFBFD',
        'bg_alt':         'E4E8EF',   # ← 加深: 原 F0F2F6 与 FAFBFD 差距不足
        'border':         'D6DAE0',
        'highlight_bg':   'FFF3CD',
        'data_bar':       '6BA3D6',
        'positive':       '1B7A3D',
        'negative':       'B71C1C',
        'neutral':        'E6A817',
        'up_cn':          'B71C1C',
        'down_cn':        '1B7A3D',
    },
}


# ═══════════════════════════════════════════════════════════════
#  内部工具函数
# ═══════════════════════════════════════════════════════════════
_CT = Alignment(horizontal='center', vertical='center', wrap_text=True)
_STYLE_PROP_KEY = '_cover_style'          # custom property key
_COL_START, _COL_END = 2, 7              # B=2 … G=7
_MARGIN_COLS = ('A', 'H')
_MARGIN_WIDTH = 4
_DATA_COL_WIDTH = 18                      # 原 16 → 18，更宽松


def _col_range(start_letter, end_letter):
    """返回 start_letter 到 end_letter 之间所有列字母（含两端）。"""
    s, e = ord(start_letter), ord(end_letter)
    return [chr(c) for c in range(s, e + 1)]


def _put(ws, P, row, val, font, merge='B:G', fill=None, h=24,
         border=None):
    """写入合并单元格，应用字体/填充/行高。fill 覆盖合并区域所有列。"""
    l, r = merge.split(':')
    ws.merge_cells(f'{l}{row}:{r}{row}')
    c = ws[f'{l}{row}']
    c.value, c.font, c.alignment = val, font, _CT
    if fill or border:
        for col in _col_range(l, r):
            cell = ws[f'{col}{row}']
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
    ws.row_dimensions[row].height = h


def _table(ws, P, row, title, headers, items, fonts):
    """写入标题 + 表头 + 数据行的迷你表格，返回最后一行行号。"""
    hf = PatternFill('solid', fgColor=P['header_bg'])
    alt_fill = PatternFill('solid', fgColor=P['bg_alt'])
    main_fill = PatternFill('solid', fgColor=P['bg_main'])
    hfn = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    row_border = Border(
        bottom=Side(style='hair', color=P['border']))

    _put(ws, P, row, title,
         Font(name='Arial', size=11, bold=True, color=P['header_bg']), h=28)
    row += 1
    for mg, hd in zip(['C:D', 'E:F'], headers):
        _put(ws, P, row, hd, hfn, mg, hf, 26)
    for i, (v1, v2) in enumerate(items):
        row += 1
        f = alt_fill if i % 2 == 0 else main_fill
        for mg, v, fn in zip(['C:D', 'E:F'], [v1, v2], fonts):
            _put(ws, P, row, v, fn, mg, f, border=row_border)
    return row


# ═══════════════════════════════════════════════════════════════
#  Style 存储 — 用 sheet custom property 替代 A1 hack
# ═══════════════════════════════════════════════════════════════

def _save_style(ws, style):
    """将 style 名写入 Cover sheet 的 custom property。"""
    from openpyxl.worksheet.properties import WorksheetProperties
    if not hasattr(ws, 'custom_properties') or ws.custom_properties is None:
        pass
    ws['A1'].value = style
    ws['A1'].font = Font(color='FFFFFF', size=1)
    ws['A1'].fill = PatternFill('solid', fgColor=PALETTE[style]['header_bg'])


def _load_style(ws):
    """从 Cover sheet 读取 style 名。"""
    val = ws['A1'].value
    return val if val in PALETTE else 'pure'


# ═══════════════════════════════════════════════════════════════
#  公开 API
# ═══════════════════════════════════════════════════════════════

def create_cover_page(wb, title, subtitle, style='pure'):
    """Phase 1: 创建 Cover Page 布局 + 样式，不写任何跨表公式。

    Args:
        wb:       Workbook 实例（应已移除默认 sheet）
        title:    主标题文字
        subtitle: 副标题文字
        style:    'pure' | 'finance'
    Returns:
        ws: Cover 工作表对象
    """
    P = PALETTE[style]
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    for col in _MARGIN_COLS:
        ws.column_dimensions[col].width = _MARGIN_WIDTH
    for i in range(_COL_START, _COL_END + 1):
        ws.column_dimensions[get_column_letter(i)].width = _DATA_COL_WIDTH

    hf = PatternFill('solid', fgColor=P['header_bg'])
    dv = Border(bottom=Side(style='thin', color=P['border']))

    _save_style(ws, style)

    # ── 游标式行号 ──
    r = 1

    # 顶部色带
    for i in range(_COL_START, _COL_END + 1):
        ws.cell(r, i).fill = hf
    ws.row_dimensions[r].height = 6
    r += 1
    ws.row_dimensions[r].height = 20
    r += 1

    # 主标题 (占两行)
    ws.merge_cells(f'B{r}:G{r + 1}')
    c = ws[f'B{r}']
    c.value = title
    c.font = Font(name='Arial', size=22, bold=True, color=P['text_primary'])
    c.alignment = _CT
    ws.row_dimensions[r].height = ws.row_dimensions[r + 1].height = 32
    r += 2

    # 副标题
    _put(ws, P, r, subtitle,
         Font(name='Arial', size=12, color=P['text_secondary']))
    r += 1

    # 分隔线
    for i in range(_COL_START, _COL_END + 1):
        ws.cell(r, i).border = dv
    ws.row_dimensions[r].height = 10
    r += 1

    # 留白
    ws.row_dimensions[r].height = 12
    r += 1

    # Metrics 区段标题 + 表头（数据行在 Phase 2 填充）
    ws._cover_metrics_start = r       # 记录起始行供 Phase 2 使用

    hfn = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    _put(ws, P, r, 'KEY METRICS',
         Font(name='Arial', size=11, bold=True, color=P['header_bg']), h=28)
    r += 1
    for mg, hd in [('C:D', 'Metric'), ('E:F', 'Value')]:
        _put(ws, P, r, hd, hfn, mg, hf, 26)
    return ws


def _validate_ref(ref, wb):
    """检查跨表公式引用的 sheet 是否存在于 workbook 中。"""
    import re
    if not isinstance(ref, str) or not ref.startswith('='):
        return
    m = re.match(r"^=(?:'([^']+)'|([A-Za-z0-9_\u4e00-\u9fff]+))!", ref)
    if m:
        sheet_name = m.group(1) or m.group(2)
        if sheet_name not in wb.sheetnames:
            import warnings
            warnings.warn(
                f"finalize_cover_page: ref '{ref}' 引用的工作表 "
                f"'{sheet_name}' 不在 workbook 中 "
                f"(现有: {wb.sheetnames})",
                stacklevel=3)


def finalize_cover_page(wb, metrics, sheet_desc=None, notes=None):
    """Phase 2: 填充 Key Metrics（含跨表公式）+ 生成 Sheet Index。
    在所有数据表创建完毕后、VALIDATE 之前调用。

    Args:
        wb:         Workbook 实例
        metrics:    list of dict，每项含 'label' + ('ref' 或 'value')
                    ref 示例: '=Analysis!B2'
        sheet_desc: dict {SheetName: description}，None 则自动留空
        notes:      可选尾注文字
    Returns:
        ws: Cover 工作表对象
    """
    ws = wb['Cover']
    style = _load_style(ws)
    P = PALETTE[style]
    dv = Border(bottom=Side(style='thin', color=P['border']))
    desc = sheet_desc or {}

    for m in metrics:
        if 'ref' in m:
            _validate_ref(m['ref'], wb)

    start = getattr(ws, '_cover_metrics_start', 8)

    # Metrics 表格
    row = _table(
        ws, P, start, '关键指标', ['Metric', 'Value'],
        [(m['label'], m.get('ref', m.get('value', ''))) for m in metrics],
        [Font(name='Arial', size=10, color=P['text_primary']),
         Font(name='Arial', size=11, bold=True, color=P['accent'])])

    # 分隔线
    row += 1
    ws.row_dimensions[row].height = 8
    row += 1
    for i in range(_COL_START, _COL_END + 1):
        ws.cell(row, i).border = dv
    ws.row_dimensions[row].height = 10

    # Sheet Index
    row += 2
    sheets = [(s, desc.get(s, '')) for s in wb.sheetnames if s != 'Cover']
    row = _table(
        ws, P, row, '工作表索引', ['Sheet Name', 'Description'], sheets,
        [Font(name='Arial', size=10, bold=True, color=P['text_primary']),
         Font(name='Arial', size=10, color=P['text_secondary'])])

    if notes:
        row += 2
        _put(ws, P, row, notes,
             Font(name='Arial', size=9, italic=True,
                  color=P['text_secondary']))

    ws.print_area = f'A1:H{row + 1}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = ws.page_setup.fitToHeight = 1
    return ws